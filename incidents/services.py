# Replace your incidents/services.py file with this corrected version

from django.db.models import Q
from datetime import datetime, timedelta
from django.utils import timezone
from .models import (
    TransportNetworkIncident, FileAccessNetworkIncident,
    RadioAccessNetworkIncident, CoreNetworkIncident,
    BackboneInternetNetworkIncident
)


class IncidentSearchService:
    """Service class for handling incident search and filtering"""
    
    def __init__(self, model_class):
        self.model_class = model_class
    
    def search_incidents(self, search_form_data):
        """
        Main search method that applies all filters and returns QuerySet
        """
        queryset = self.model_class.objects.select_related('created_by', 'updated_by')
        
        # Apply text search
        if search_form_data.get('search_query'):
            queryset = self._apply_text_search(queryset, search_form_data['search_query'])
        
        # Apply date range filters
        if search_form_data.get('date_from'):
            queryset = queryset.filter(date_time_incident__gte=search_form_data['date_from'])
        
        if search_form_data.get('date_to'):
            queryset = queryset.filter(date_time_incident__lte=search_form_data['date_to'])
        
        # Apply status filter
        if search_form_data.get('status'):
            queryset = self._apply_status_filter(queryset, search_form_data['status'])
        
        # Apply cause and origin filters
        if search_form_data.get('cause'):
            queryset = queryset.filter(cause=search_form_data['cause'])
        
        if search_form_data.get('origin'):
            queryset = queryset.filter(origin=search_form_data['origin'])
        
        # Apply network-specific filters
        queryset = self._apply_network_specific_filters(queryset, search_form_data)
        
        # Apply sorting
        sort_by = search_form_data.get('sort_by', '-date_time_incident')
        queryset = queryset.order_by(sort_by)
        
        return queryset
    
    def _apply_text_search(self, queryset, search_query):
        """Apply text search across relevant fields"""
        search_query = search_query.strip()
        
        if not search_query:
            return queryset
        
        # Base fields common to all incidents
        q_objects = Q()
        q_objects |= Q(id__icontains=search_query)
        q_objects |= Q(cause__icontains=search_query)
        q_objects |= Q(origin__icontains=search_query)
        q_objects |= Q(impact_comment__icontains=search_query)
        q_objects |= Q(created_by__username__icontains=search_query)
        
        # Network-specific search fields - CORRECTED field names
        if self.model_class == TransportNetworkIncident:
            q_objects |= Q(region_loop__icontains=search_query)
            q_objects |= Q(system_capacity__icontains=search_query)
            q_objects |= Q(extremity_a__icontains=search_query)
            q_objects |= Q(extremity_b__icontains=search_query)
        
        elif self.model_class == FileAccessNetworkIncident:
            q_objects |= Q(do_wilaya__icontains=search_query)  # CORRECTED: do_wilaya not wilaya
            q_objects |= Q(zone_metro__icontains=search_query)
            q_objects |= Q(site__icontains=search_query)
            q_objects |= Q(ip_address__icontains=search_query)
        
        elif self.model_class == RadioAccessNetworkIncident:
            q_objects |= Q(do_wilaya__icontains=search_query)  # CORRECTED: do_wilaya not wilaya
            q_objects |= Q(site__icontains=search_query)
            q_objects |= Q(ip_address__icontains=search_query)
        
        elif self.model_class == CoreNetworkIncident:
            q_objects |= Q(platform__icontains=search_query)
            q_objects |= Q(region_node__icontains=search_query)
            q_objects |= Q(site__icontains=search_query)
            q_objects |= Q(extremity_a__icontains=search_query)
            q_objects |= Q(extremity_b__icontains=search_query)
        
        elif self.model_class == BackboneInternetNetworkIncident:
            q_objects |= Q(interconnect_type__icontains=search_query)
            q_objects |= Q(platform_igw__icontains=search_query)
            q_objects |= Q(link_label__icontains=search_query)
        
        return queryset.filter(q_objects)
    
    def _apply_status_filter(self, queryset, status):
        """Apply status-based filtering"""
        now = timezone.now()
        
        if status == 'active':
            return queryset.filter(date_time_recovery__isnull=True)
        
        elif status == 'resolved':
            return queryset.filter(date_time_recovery__isnull=False)
        
        elif status == 'new':
            # Active incidents less than 1 hour old
            one_hour_ago = now - timedelta(hours=1)
            return queryset.filter(
                date_time_recovery__isnull=True,
                date_time_incident__gte=one_hour_ago
            )
        
        elif status == 'low':
            # Active incidents 1-2 hours old
            one_hour_ago = now - timedelta(hours=1)
            two_hours_ago = now - timedelta(hours=2)
            return queryset.filter(
                date_time_recovery__isnull=True,
                date_time_incident__lte=one_hour_ago,
                date_time_incident__gte=two_hours_ago
            )
        
        elif status == 'medium':
            # Active incidents 2-4 hours old
            two_hours_ago = now - timedelta(hours=2)
            four_hours_ago = now - timedelta(hours=4)
            return queryset.filter(
                date_time_recovery__isnull=True,
                date_time_incident__lte=two_hours_ago,
                date_time_incident__gte=four_hours_ago
            )
        
        elif status == 'critical':
            # Active incidents more than 4 hours old
            four_hours_ago = now - timedelta(hours=4)
            return queryset.filter(
                date_time_recovery__isnull=True,
                date_time_incident__lte=four_hours_ago
            )
        
        return queryset
    
    def _apply_network_specific_filters(self, queryset, search_form_data):
        """Apply network-specific filters based on model type - CORRECTED field names"""
        
        if self.model_class == TransportNetworkIncident:
            if search_form_data.get('region_loop'):
                queryset = queryset.filter(region_loop=search_form_data['region_loop'])
            
            if search_form_data.get('system_capacity'):
                queryset = queryset.filter(system_capacity=search_form_data['system_capacity'])
            
            if search_form_data.get('extremity_a'):
                queryset = queryset.filter(extremity_a__icontains=search_form_data['extremity_a'])
            
            if search_form_data.get('extremity_b'):
                queryset = queryset.filter(extremity_b__icontains=search_form_data['extremity_b'])
        
        elif self.model_class in [FileAccessNetworkIncident, RadioAccessNetworkIncident]:
            if search_form_data.get('do_wilaya'):  # CORRECTED: do_wilaya not wilaya
                queryset = queryset.filter(do_wilaya=search_form_data['do_wilaya'])
            
            if search_form_data.get('site'):
                queryset = queryset.filter(site__icontains=search_form_data['site'])
            
            if search_form_data.get('ip_address'):
                queryset = queryset.filter(ip_address__icontains=search_form_data['ip_address'])
            
            # File Access specific
            if (self.model_class == FileAccessNetworkIncident and 
                search_form_data.get('zone_metro')):
                queryset = queryset.filter(zone_metro__icontains=search_form_data['zone_metro'])
        
        elif self.model_class == CoreNetworkIncident:
            if search_form_data.get('platform'):
                queryset = queryset.filter(platform=search_form_data['platform'])
            
            if search_form_data.get('region_node'):
                queryset = queryset.filter(region_node=search_form_data['region_node'])
            
            if search_form_data.get('site'):
                queryset = queryset.filter(site__icontains=search_form_data['site'])
        
        elif self.model_class == BackboneInternetNetworkIncident:
            if search_form_data.get('interconnect_type'):
                queryset = queryset.filter(interconnect_type=search_form_data['interconnect_type'])
            
            if search_form_data.get('platform_igw'):
                queryset = queryset.filter(platform_igw=search_form_data['platform_igw'])
            
            if search_form_data.get('link_label'):
                queryset = queryset.filter(link_label__icontains=search_form_data['link_label'])
        
        return queryset
    
    def get_search_statistics(self, original_queryset, filtered_queryset):
        """Generate search statistics for display"""
        return {
            'total_incidents': original_queryset.count(),
            'filtered_incidents': filtered_queryset.count(),
            'active_incidents': filtered_queryset.filter(date_time_recovery__isnull=True).count(),
            'resolved_incidents': filtered_queryset.filter(date_time_recovery__isnull=False).count(),
        }
    
    def get_optimized_statistics(self, original_queryset, filtered_queryset):
        """Generate optimized search statistics for display"""
        # Use efficient count queries
        original_count = original_queryset.count()
        filtered_count = filtered_queryset.count()
    
        # Efficient active/resolved counts on filtered set
        active_count = filtered_queryset.filter(date_time_recovery__isnull=True).count()
        resolved_count = filtered_count - active_count
    
        return {
        'total_incidents': original_count,
        'filtered_incidents': filtered_count,
        'active_incidents': active_count,
        'resolved_incidents': resolved_count,
        }
    
    def get_bulk_incident_data(self, queryset, limit=None):
        """
        Optimized method for bulk incident data retrieval with minimal database hits
        """
        if limit:
            queryset = queryset[:limit]
        
        # Use values() to get only needed fields, reducing memory usage
        incident_data = queryset.values(
            'id', 'date_time_incident', 'date_time_recovery', 'duration_minutes',
            'cause', 'origin', 'impact_comment', 'is_resolved', 'is_archived',
            'created_by__username', 'updated_by__username'
        )
        
        return list(incident_data)

    def get_incident_summary_stats(self, queryset):
        """
        Get comprehensive statistics with a single database query
        """
        from django.db.models import Count, Q
        
        stats = queryset.aggregate(
            total_count=Count('id'),
            active_count=Count('id', filter=Q(date_time_recovery__isnull=True)),
            resolved_count=Count('id', filter=Q(date_time_recovery__isnull=False)),
            # Add severity-based counts
            new_count=Count('id', filter=Q(
                date_time_recovery__isnull=True,
                date_time_incident__gte=timezone.now() - timedelta(hours=1)
            )),
            critical_count=Count('id', filter=Q(
                date_time_recovery__isnull=True,
                date_time_incident__lte=timezone.now() - timedelta(hours=4)
            ))
        )
        
        return stats

# Factory function to get appropriate search service - CORRECTED network type mapping
def get_search_service(network_type):
    """Factory function to return appropriate search service"""
    
    # CORRECTED: Match your views.py NETWORKS dictionary keys
    model_mapping = {
        'transport': TransportNetworkIncident,
        'file_access': FileAccessNetworkIncident,  # CORRECTED: file_access not file-access
        'radio_access': RadioAccessNetworkIncident,  # CORRECTED: radio_access not radio-access
        'core': CoreNetworkIncident,
        'backbone_internet': BackboneInternetNetworkIncident,  # CORRECTED: backbone_internet not backbone-internet
    }
    
    model_class = model_mapping.get(network_type)
    if not model_class:
        raise ValueError(f"Unknown network type: {network_type}")
    
    return IncidentSearchService(model_class)

# ============================================================================
# CSV/EXCEL EXPORT SERVICE (Task 2: Phase 4)
# ============================================================================

import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


class IncidentExportService:
    """Service class for exporting incidents to CSV/Excel formats"""
    
    def __init__(self, queryset, network_type):
        self.queryset = queryset
        self.network_type = network_type
        self.network_name = self._get_network_name()
    
    def _get_network_name(self):
        """Get display name for network type"""
        network_names = {
            'transport': 'Transport Networks',
            'file_access': 'File Access Networks',
            'radio_access': 'Radio Access Networks',
            'core': 'Core Networks',
            'backbone_internet': 'Backbone Internet Networks',
        }
        return network_names.get(self.network_type, 'Unknown Network')
    
    def export_to_csv(self):
        """Export incidents to CSV format"""
        output = io.StringIO()
        
        # Get field configuration for this network type
        headers, field_getters = self._get_export_fields()
        
        writer = csv.writer(output)
        
        # Write header row
        writer.writerow(headers)
        
        # Write data rows
        for incident in self.queryset:
            row = [getter(incident) for getter in field_getters]
            writer.writerow(row)
        
        return output.getvalue()
    
    def export_to_excel(self):
        """Export incidents to Excel format with formatting"""
        wb = Workbook()
        ws = wb.active
        ws.title = self.network_type.replace('_', ' ').title()[:31]  # Excel sheet name limit
        
        # Get field configuration
        headers, field_getters = self._get_export_fields()
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="003d7a", end_color="003d7a", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        cell_border = Border(
            left=Side(style='thin', color="CCCCCC"),
            right=Side(style='thin', color="CCCCCC"),
            top=Side(style='thin', color="CCCCCC"),
            bottom=Side(style='thin', color="CCCCCC")
        )
        
        # Write header row with styling
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border
        
        # Write data rows with color coding
        for row_num, incident in enumerate(self.queryset, 2):
            # Get severity class for color coding
            severity_class = incident.get_severity_class()
            row_fill = self._get_severity_fill(severity_class)
            
            for col_num, getter in enumerate(field_getters, 1):
                value = getter(incident)
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = cell_border
                
                # Apply color coding to first column
                if col_num == 1:
                    cell.fill = row_fill
        
        # Auto-size columns
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(headers[col_num - 1]))
            
            for row in ws.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
                try:
                    cell_value = str(row[0].value) if row[0].value else ""
                    max_length = max(max_length, len(cell_value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _get_severity_fill(self, severity_class):
        """Get Excel fill color based on severity"""
        severity_colors = {
            'incident-new': PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"),
            'incident-low': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            'incident-medium': PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid"),
            'incident-critical': PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
            'incident-resolved': PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        }
        return severity_colors.get(severity_class, PatternFill())
    
    def _get_export_fields(self):
        """Get headers and field getter functions for each network type"""
        
        # Common fields for all networks
        common_headers = [
            'Incident ID',
            'Start Date/Time',
            'Recovery Date/Time',
            'Duration',
            'Status',
            'Severity',
            'Cause',
            'Origin',
            'Impact/Comment',
            'Created By',
            'Created At',
        ]
        
        common_getters = [
            lambda i: str(i.id)[:8] + '...',
            lambda i: i.date_time_incident.strftime('%Y-%m-%d %H:%M:%S') if i.date_time_incident else '',
            lambda i: i.date_time_recovery.strftime('%Y-%m-%d %H:%M:%S') if i.date_time_recovery else '',
            lambda i: i.get_duration_display(),
            lambda i: 'Resolved' if i.is_resolved else 'Active',
            lambda i: i.get_severity_display(),
            lambda i: i.cause or '',
            lambda i: i.origin or '',
            lambda i: i.impact_comment or '',
            lambda i: i.created_by.username if i.created_by else '',
            lambda i: i.created_at.strftime('%Y-%m-%d %H:%M:%S') if i.created_at else '',
        ]
        
        # Network-specific fields
        if self.network_type == 'transport':
            network_headers = ['Region/Loop', 'System/Capacity', 'Extremity A', 'Extremity B', 'Responsibility']
            network_getters = [
                lambda i: i.region_loop or '',
                lambda i: i.system_capacity or '',
                lambda i: i.extremity_a or '',
                lambda i: i.extremity_b or '',
                lambda i: i.responsibility or '',
            ]
        
        elif self.network_type == 'file_access':
            network_headers = ['DO/Wilaya', 'Zone/Metro', 'Site', 'IP Address']
            network_getters = [
                lambda i: i.do_wilaya or '',
                lambda i: i.zone_metro or '',
                lambda i: i.site or '',
                lambda i: i.ip_address or '',
            ]
        
        elif self.network_type == 'radio_access':
            network_headers = ['DO/Wilaya', 'Site', 'IP Address']
            network_getters = [
                lambda i: i.do_wilaya or '',
                lambda i: i.site or '',
                lambda i: i.ip_address or '',
            ]
        
        elif self.network_type == 'core':
            network_headers = ['Platform', 'Region/Node', 'Site', 'Extremity A', 'Extremity B']
            network_getters = [
                lambda i: i.platform or '',
                lambda i: i.region_node or '',
                lambda i: i.site or '',
                lambda i: i.extremity_a or '',
                lambda i: i.extremity_b or '',
            ]
        
        elif self.network_type == 'backbone_internet':
            network_headers = ['Interconnect Type', 'Platform/IGW', 'Link Label']
            network_getters = [
                lambda i: i.interconnect_type or '',
                lambda i: i.platform_igw or '',
                lambda i: i.link_label or '',
            ]
        
        else:
            network_headers = []
            network_getters = []
        
        # Combine headers and getters
        headers = network_headers + common_headers
        field_getters = network_getters + common_getters
        
        return headers, field_getters
    
    def get_filename(self, format='csv'):
        """Generate appropriate filename for export"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        network_slug = self.network_type.replace('_', '-')
        extension = 'csv' if format == 'csv' else 'xlsx'
        
        return f'incidents_{network_slug}_{timestamp}.{extension}'


def get_export_service(queryset, network_type):
    """Factory function to get export service"""
    return IncidentExportService(queryset, network_type)

