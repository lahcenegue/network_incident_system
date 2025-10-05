from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator, EmptyPage, InvalidPage
from django.utils import timezone
from django.db.models import Q
from datetime import timedelta
from .services import get_search_service
from .forms import get_search_form_class
from .models import (
    TransportNetworkIncident, FileAccessNetworkIncident, 
    RadioAccessNetworkIncident, CoreNetworkIncident, 
    BackboneInternetNetworkIncident, DropdownConfiguration
)
from .forms import (
    TransportNetworkIncidentForm, FileAccessNetworkIncidentForm,
    RadioAccessNetworkIncidentForm, CoreNetworkIncidentForm,
    BackboneInternetNetworkIncidentForm, get_incident_form_class, update_form_common_fields
)
from .utils import get_incident_color_class
import json

def format_datetime_for_input(dt):
    """
    Format datetime object for HTML5 datetime-local input.
    Required format: YYYY-MM-DDTHH:MM
    """
    if not dt:
        return None
    
    # Convert to local timezone if aware
    if timezone.is_aware(dt):
        dt = timezone.localtime(dt)
    
    # Format as HTML5 datetime-local expects: YYYY-MM-DDTHH:MM
    return dt.strftime('%Y-%m-%dT%H:%M')

# Fixed Network configuration - corrected backbone_internet naming
NETWORKS = {
    'transport': {
        'name': 'Transport Networks',
        'model': TransportNetworkIncident,
        'form': TransportNetworkIncidentForm,
        'template': 'incidents/transport_networks.html'
    },
    'file_access': {
        'name': 'File Access Networks',
        'model': FileAccessNetworkIncident,
        'form': FileAccessNetworkIncidentForm,
        'template': 'incidents/file_access_networks.html'
    },
    'radio_access': {
        'name': 'Radio Access Networks',
        'model': RadioAccessNetworkIncident,
        'form': RadioAccessNetworkIncidentForm,
        'template': 'incidents/radio_access_networks.html'
    },
    'core': {
        'name': 'Core Networks',
        'model': CoreNetworkIncident,
        'form': CoreNetworkIncidentForm,
        'template': 'incidents/core_networks.html'
    },
    'backbone_internet': {  # Fixed: was missing _internet
        'name': 'Backbone Internet Networks',
        'model': BackboneInternetNetworkIncident,
        'form': BackboneInternetNetworkIncidentForm,
        'template': 'incidents/backbone_internet_networks.html'
    }
}

@login_required
def network_incidents_view(request, network_type):
    """Enhanced view with advanced pagination for large datasets"""
    if network_type not in NETWORKS:
        messages.error(request, f"Invalid network type: {network_type}")
        return redirect('dashboard:dashboard')
    
    network_config = NETWORKS[network_type]
    model = network_config['model']
    
    try:
        # Get search service and form
        search_service = get_search_service(network_type)
        search_form_class = get_search_form_class(network_type)
        search_form = search_form_class(request.GET) if search_form_class else None
        
        # OPTIMIZED: Use select_related and only() for memory efficiency
        base_queryset = model.objects.filter(is_archived=False).select_related('created_by', 'updated_by').only(
            'id', 'date_time_incident', 'date_time_recovery', 'duration_minutes',
            'cause', 'origin', 'is_resolved', 'created_by__username', 'updated_by__username',
            # Network-specific essential fields
            *get_essential_fields_for_network(network_type)
        )
        
        # Apply search filters
        filtered_queryset = base_queryset
        search_active = bool(False)
        
        if search_form and search_form.is_valid():
            form_data = search_form.cleaned_data
            search_active = any(value for value in form_data.values() if value)
            
            if search_active:
                filtered_queryset = search_service.search_incidents(form_data)
            else:
                filtered_queryset = base_queryset.order_by('-date_time_incident')
        else:
            filtered_queryset = base_queryset.order_by('-date_time_incident')
        
        # ENHANCED: Dynamic page size with memory limits
        page_size = min(int(request.GET.get('size', 25)), 100)  # Max 100 per page
        paginator = Paginator(filtered_queryset, page_size)
        page_number = request.GET.get('page', 1)
        
        try:
            incidents = paginator.page(page_number)
        except (EmptyPage, InvalidPage):
            incidents = paginator.page(1)
        
        # OPTIMIZED: Get statistics with efficient queries
        search_stats = search_service.get_optimized_statistics(base_queryset, filtered_queryset)
        
        # MEMORY OPTIMIZED: Limit recent resolved to prevent memory issues
        recent_resolved_qs = filtered_queryset.filter(
            is_archived=False,
            date_time_recovery__isnull=False,
            date_time_recovery__gte=timezone.now() - timedelta(hours=24)
        ).order_by('-date_time_recovery')[:10]
        
        # MEMORY OPTIMIZED: Get active incidents with limits
        active_incidents_qs = filtered_queryset.filter(
            is_archived=False,
            date_time_recovery__isnull=True
        ).order_by('-date_time_incident')
        
        # Calculate severity counts efficiently
        active_incidents_list = list(active_incidents_qs[:100])  # Limit to first 100 for stats
        severity_counts = {'new': 0, 'low': 0, 'medium': 0, 'critical': 0}

        for incident in active_incidents_list:
            severity = incident.get_severity_class()
            if severity == 'incident-new':
                severity_counts['new'] += 1
            elif severity == 'incident-low':
                severity_counts['low'] += 1
            elif severity == 'incident-medium':
                severity_counts['medium'] += 1
            elif severity == 'incident-critical':
                severity_counts['critical'] += 1
        
        context = {
            'network_type': network_type,
            'network_name': network_config['name'],
            'incidents': incidents,
            'search_form': search_form,
            'search_active': search_active,
            'search_stats': search_stats,
            'total_incidents': search_stats['total_incidents'],
            'active_incidents': active_incidents_qs,
            'resolved_incidents': search_stats['resolved_incidents'],
            'resolved_recent': recent_resolved_qs,
            'severity_counts': severity_counts,
            'page_obj': incidents,  # For pagination template
        }
        
        return render(request, network_config['template'], context)
        
    except Exception as e:
        messages.error(request, f"Error loading incidents: {str(e)}")
        return render(request, network_config['template'], {
            'network_type': network_type,
            'network_name': network_config['name'],
            'incidents': [],
            'search_form': None,
            'search_active': False,
            'total_incidents': 0,
            'active_incidents': [],
            'resolved_incidents': 0,
            'resolved_recent': [],
            'severity_counts': {'new': 0, 'low': 0, 'medium': 0, 'critical': 0},
            'error': str(e)
        })

# NEW: Helper function for memory optimization
def get_essential_fields_for_network(network_type):
    """Return essential fields for each network type to optimize queries"""
    field_map = {
        'transport': ['region_loop', 'system_capacity', 'extremity_a', 'extremity_b'],
        'file_access': ['do_wilaya', 'zone_metro', 'site', 'ip_address'],
        'radio_access': ['do_wilaya', 'site', 'ip_address'],
        'core': ['platform', 'region_node', 'site', 'extremity_a', 'extremity_b'],
        'backbone_internet': ['interconnect_type', 'platform_igw', 'link_label'],
    }
    return field_map.get(network_type, [])

@login_required
@require_http_methods(["GET", "POST"])
def add_incident_view(request, network_type):
    """Enhanced view to add new incidents with form data preservation"""
    
    # Validate network type first
    if network_type not in NETWORKS:
        messages.error(request, f"Invalid network type: {network_type}")
        return redirect('dashboard:dashboard')
    
    # Get dropdown data directly from database
    from .models import DropdownConfiguration
    
    # Initialize form data with POST data if available, empty dict otherwise
    form_data = request.POST if request.method == 'POST' else {}
    
    context = {
        'network_type': NETWORKS[network_type]['name'],
        'action': 'Add New',
        'form_data': form_data,  # Pass form data to template
        # Common dropdown choices for all networks
        'cause_choices': DropdownConfiguration.objects.filter(
            category='cause', is_active=True
        ).order_by('sort_order', 'value'),
        'origin_choices': DropdownConfiguration.objects.filter(
            category='origin', is_active=True
        ).order_by('sort_order', 'value'),
    }
    
    # Add network-specific choices
    if network_type == 'transport':
        context.update({
            'region_choices': DropdownConfiguration.objects.filter(
                category='region_loop', is_active=True
            ).order_by('sort_order', 'value'),
            'system_choices': DropdownConfiguration.objects.filter(
                category='system_capacity', is_active=True
            ).order_by('sort_order', 'value'),
            'dot_choices': DropdownConfiguration.objects.filter(
                category='dot_states', is_active=True
            ).order_by('sort_order', 'value'),
        })
    elif network_type in ['file_access', 'radio_access']:
        context.update({
            'wilaya_choices': DropdownConfiguration.objects.filter(
                category='wilayas', is_active=True
            ).order_by('sort_order', 'value'),
        })
    elif network_type == 'core':
        context.update({
            'platform_choices': DropdownConfiguration.objects.filter(
                category='platforms', is_active=True
            ).order_by('sort_order', 'value'),
            'region_node_choices': DropdownConfiguration.objects.filter(
                category='region_nodes', is_active=True
            ).order_by('sort_order', 'value'),
            'dot_choices': DropdownConfiguration.objects.filter(
                category='dot_states', is_active=True
            ).order_by('sort_order', 'value'),
        })
    elif network_type == 'backbone_internet':
        context.update({
            'interconnect_choices': DropdownConfiguration.objects.filter(
                category='interconnect_types', is_active=True
            ).order_by('sort_order', 'value'),
            'platform_igw_choices': DropdownConfiguration.objects.filter(
                category='platform_igws', is_active=True
            ).order_by('sort_order', 'value'),
        })
    
    if request.method == 'POST':
        try:
            # Get the form class
            form_class = get_incident_form_class(network_type)
            form = form_class(request.POST, user=request.user)
            
            if form.is_valid():
                incident = form.save(commit=False)
                incident.created_by = request.user
                incident.save()
                
                messages.success(
                    request, 
                    f"Incident {str(incident.id)[:8]} created successfully for {NETWORKS[network_type]['name']}"
                )
                
                # Redirect to the appropriate network incidents list
                redirect_urls = {
                    'transport': 'incidents:transport_incidents',
                    'file_access': 'incidents:file_access_incidents',
                    'radio_access': 'incidents:radio_access_incidents',
                    'core': 'incidents:core_incidents',
                    'backbone_internet': 'incidents:backbone_internet_incidents'
                }
                
                redirect_url = redirect_urls.get(network_type, 'dashboard:dashboard')
                return redirect(redirect_url)
                
            else:
                # Handle form errors but preserve data
                for field, errors in form.errors.items():
                    for error in errors:
                        if field == '__all__':
                            messages.error(request, error)
                        else:
                            field_label = field.replace('_', ' ').title()
                            messages.error(request, f"{field_label}: {error}")
                
                # Form data is already in context, template will preserve it
                
        except Exception as e:
            messages.error(request, f"Error creating incident: {str(e)}")
    
    return render(request, 'incidents/incident_form.html', context)

@login_required
@require_http_methods(["GET", "POST"])
def edit_incident_view(request, network_type, incident_id):
    """Enhanced view to edit existing incidents with validation"""
    if network_type not in NETWORKS:
        messages.error(request, f"Invalid network type: {network_type}")
        return redirect('dashboard:dashboard')
    
    network_config = NETWORKS[network_type]
    model = network_config['model']
    form_class = network_config['form']
    
    try:
        incident = get_object_or_404(model, id=incident_id)

        from .models import DropdownConfiguration
        
        dropdown_context = {
            'cause_choices': list(DropdownConfiguration.objects.filter(
                category='cause', is_active=True
            ).order_by('sort_order', 'value')),
            'origin_choices': list(DropdownConfiguration.objects.filter(
                category='origin', is_active=True
            ).order_by('sort_order', 'value')),
        }
        
        # Add network-specific dropdown choices
        if network_type == 'transport':
            dropdown_context.update({
                'region_choices': list(DropdownConfiguration.objects.filter(
                    category='region_loop', is_active=True
                ).order_by('sort_order', 'value')),
                'system_choices': list(DropdownConfiguration.objects.filter(
                    category='system_capacity', is_active=True
                ).order_by('sort_order', 'value')),
                'dot_choices': list(DropdownConfiguration.objects.filter(
                    category='dot_states', is_active=True
                ).order_by('sort_order', 'value')),
            })
        elif network_type in ['file_access', 'radio_access']:
            dropdown_context['wilaya_choices'] = list(DropdownConfiguration.objects.filter(
                category='wilayas', is_active=True
            ).order_by('sort_order', 'value'))
        elif network_type == 'core':
            dropdown_context.update({
                'platform_choices': list(DropdownConfiguration.objects.filter(
                    category='platforms', is_active=True
                ).order_by('sort_order', 'value')),
                'region_node_choices': list(DropdownConfiguration.objects.filter(
                    category='region_nodes', is_active=True
                ).order_by('sort_order', 'value')),
                'dot_choices': list(DropdownConfiguration.objects.filter(
                    category='dot_states', is_active=True
                ).order_by('sort_order', 'value')),
            })
        elif network_type == 'backbone_internet':
            dropdown_context.update({
                'interconnect_choices': list(DropdownConfiguration.objects.filter(
                    category='interconnect_types', is_active=True
                ).order_by('sort_order', 'value')),
                'platform_igw_choices': list(DropdownConfiguration.objects.filter(
                    category='platform_igws', is_active=True
                ).order_by('sort_order', 'value')),
            })
        
        # Check if user can edit this incident
        if not request.user.is_admin() and incident.created_by != request.user:
            messages.error(request, "You can only edit incidents you created.")
            return redirect(f'incidents:{network_type}_incidents')
        
        if request.method == 'POST':
            form = form_class(request.POST, instance=incident, user=request.user)
            form = update_form_common_fields(form, network_type)
            
            if form.is_valid():
                updated_incident = form.save(commit=False)
                updated_incident.updated_by = request.user
                updated_incident.updated_at = timezone.now()
                updated_incident.save()
                
                messages.success(
                    request, 
                    f"Incident {str(incident.id)[:8]} updated successfully"
                )
                
                return redirect(f'incidents:{network_type}_incidents')
            else:
                # Collect and display form errors
                for field, errors in form.errors.items():
                    for error in errors:
                        if field == '__all__':
                            messages.error(request, error)
                        else:
                            field_label = form.fields.get(field, {}).label or field.replace('_', ' ').title()
                            messages.error(request, f"{field_label}: {error}")
        else:
            # GET request - Prepare initial data with properly formatted datetime
            initial_data = {}
            
            # CRITICAL FIX: Format datetime fields for HTML5 datetime-local input
            initial_data['date_time_incident'] = format_datetime_for_input(incident.date_time_incident)
            
            if incident.date_time_recovery:
                initial_data['date_time_recovery'] = format_datetime_for_input(incident.date_time_recovery)
            
            # Add all other fields based on network type
            if network_type == 'transport':
                initial_data.update({
                    'region_loop': incident.region_loop,
                    'system_capacity': incident.system_capacity,
                    'dot_extremity_a': incident.dot_extremity_a,
                    'extremity_a': incident.extremity_a,
                    'dot_extremity_b': incident.dot_extremity_b,
                    'extremity_b': incident.extremity_b,
                    'responsibility': incident.responsibility,
                })
            elif network_type == 'file_access':
                initial_data.update({
                    'do_wilaya': incident.do_wilaya,
                    'zone_metro': incident.zone_metro,
                    'site': incident.site,
                    'ip_address': incident.ip_address,
                })
            elif network_type == 'radio_access':
                initial_data.update({
                    'do_wilaya': incident.do_wilaya,
                    'site': incident.site,
                    'ip_address': incident.ip_address,
                })
            elif network_type == 'core':
                initial_data.update({
                    'platform': incident.platform,
                    'region_node': incident.region_node,
                    'site': incident.site,
                    'dot_extremity_a': incident.dot_extremity_a,
                    'extremity_a': incident.extremity_a,
                    'dot_extremity_b': incident.dot_extremity_b,
                    'extremity_b': incident.extremity_b,
                })
            elif network_type == 'backbone_internet':
                initial_data.update({
                    'interconnect_type': incident.interconnect_type,
                    'platform_igw': incident.platform_igw,
                    'link_label': incident.link_label,
                })
            
            # Common fields for all networks
            initial_data.update({
                'cause': incident.cause,
                'origin': incident.origin,
                'impact_comment': incident.impact_comment,
            })
            
            # Create form with initial data AND instance
            form = form_class(initial=initial_data, instance=incident, user=request.user)
            form = update_form_common_fields(form, network_type)
        
        context = {
            'form': form,
            'incident': incident,
            'network_type': network_config['name'],
            'form_title': f'Edit {network_config["name"]} Incident',
            'submit_text': 'Update Incident',
            'cancel_url': reverse(f'incidents:{network_type}_incidents'),
            'is_edit': True,
            'action': 'Edit',
            **dropdown_context
        }
        
        return render(request, 'incidents/incident_form.html', context)
        
    except Exception as e:
        messages.error(request, f"Error editing incident: {str(e)}")
        return redirect(f'incidents:{network_type}_incidents')
    

@login_required
def incident_notification_prompt(request, network_type, incident_id):
    """Prompt user to send notifications after creating incident"""
    if network_type not in NETWORKS:
        return redirect('dashboard:dashboard')
    
    network_config = NETWORKS[network_type]
    model = network_config['model']
    
    try:
        incident = get_object_or_404(model, id=incident_id)
        
        context = {
            'incident': incident,
            'network_type': network_type,
            'network_name': network_config['name'],
        }
        
        return render(request, 'incidents/notification_prompt.html', context)
        
    except Exception as e:
        messages.error(request, f"Error loading notification prompt: {str(e)}")
        return redirect(f'incidents:{network_type}_incidents')

@login_required
def historical_incidents_view(request, network_type):
    """Enhanced view for historical incidents"""
    if network_type not in NETWORKS:
        messages.error(request, f"Invalid network type: {network_type}")
        return redirect('dashboard:dashboard')
    
    network_config = NETWORKS[network_type]
    model = network_config['model']
    
    try:
        # Get resolved incidents (those with recovery time)
        incidents_queryset = model.objects.filter(
            date_time_recovery__isnull=False
        ).select_related('created_by', 'updated_by').order_by('-date_time_recovery')
        
        # Pagination
        paginator = Paginator(incidents_queryset, 25)  # 25 historical incidents per page
        page_number = request.GET.get('page')
        incidents = paginator.get_page(page_number)
        
        # Add color class (should all be green for resolved)
        for incident in incidents:
            incident.color_class = get_incident_color_class(incident)
        
        context = {
            'network_type': network_type,
            'network_name': network_config['name'],
            'incidents': incidents,
            'total_resolved': incidents_queryset.count(),
            'page_obj': incidents,
        }
        
        return render(request, 'incidents/historical_incidents.html', context)
        
    except Exception as e:
        messages.error(request, f"Error loading historical incidents: {str(e)}")
        return render(request, 'incidents/historical_incidents.html', {
            'network_type': network_type,
            'network_name': network_config['name'],
            'incidents': [],
            'total_resolved': 0,
            'error': str(e)
        })

@login_required
@require_http_methods(["POST"])
def validate_incident_field(request, network_type):
    """AJAX endpoint for real-time field validation"""
    if network_type not in NETWORKS:
        return JsonResponse({'valid': False, 'error': 'Invalid network type'})
    
    field_name = request.POST.get('field_name')
    field_value = request.POST.get('field_value')
    incident_id = request.POST.get('incident_id')  # For edit validation
    
    try:
        form_class = NETWORKS[network_type]['form']
        
        # Create a temporary form instance for validation
        form_data = {field_name: field_value}
        form = form_class(form_data, user=request.user)
        
        # Validate specific field
        form.is_valid()  # This populates form.errors
        
        if field_name in form.errors:
            return JsonResponse({
                'valid': False, 
                'errors': form.errors[field_name]
            })
        else:
            return JsonResponse({'valid': True})
            
    except Exception as e:
        return JsonResponse({'valid': False, 'error': str(e)})
    

@login_required
@require_http_methods(["GET"])
def ajax_search_incidents(request, network_type):
    """AJAX endpoint for real-time search suggestions"""
    if network_type not in NETWORKS:
        return JsonResponse({'error': 'Invalid network type'}, status=400)
    
    search_query = request.GET.get('q', '').strip()
    
    if len(search_query) < 2:
        return JsonResponse({'suggestions': []})
    
    try:
        search_service = get_search_service(network_type)
        model = NETWORKS[network_type]['model']
        
        # Quick search for suggestions
        queryset = model.objects.filter(
            Q(id__icontains=search_query) |
            Q(impact_comment__icontains=search_query) |
            Q(cause__icontains=search_query) |
            Q(origin__icontains=search_query)
        ).select_related('created_by')[:5]  # Limit to 5 suggestions
        
        suggestions = []
        for incident in queryset:
            suggestions.append({
                'id': str(incident.id),
                'text': f"ID: {str(incident.id)[:8]}... - {incident.cause or 'Unknown cause'}",
                'url': f"/incidents/{network_type}/"
            })
        
        return JsonResponse({'suggestions': suggestions})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    

@login_required
def get_incident_detail(request, network_type, incident_id):
    """
    AJAX endpoint to fetch incident details for modal display.
    Returns HTML content as JSON response.
    """
    try:
        # Map network type to model
        model_map = {
            'transport': TransportNetworkIncident,
            'file_access': FileAccessNetworkIncident,
            'radio_access': RadioAccessNetworkIncident,
            'core': CoreNetworkIncident,
            'backbone_internet': BackboneInternetNetworkIncident,
        }
        
        # Validate network type
        if network_type not in model_map:
            return JsonResponse({
                'success': False,
                'error': 'Invalid network type'
            }, status=400)
        
        # Get the incident
        model = model_map[network_type]
        incident = model.objects.select_related('created_by', 'updated_by').get(id=incident_id)
        
        # Prepare context data
        context = {
            'incident': incident,
            'network_type': network_type,
        }
        
        # CORRECTED PATH: Use incident_management/templates/incidents/detail_sections/
        template_name = f'incidents/detail_sections/{network_type}_detail.html'
        html_content = render_to_string(template_name, context, request=request)
        
        # Return success response with HTML
        return JsonResponse({
            'success': True,
            'html': html_content,
            'edit_url': f'/incidents/{network_type}-networks/edit/{incident_id}/',
            'incident_id': str(incident_id),
        })
        
    except model.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Incident not found'
        }, status=404)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    

    
@login_required
def unified_historical_incidents_view(request):
    """
    Unified view for all archived incidents across all 5 network types.
    Shows archived incidents with filtering, search, and restore capabilities.
    """
    try:
        # Get all archived incidents from all networks
        transport_archived = TransportNetworkIncident.objects.filter(
            is_archived=True
        ).select_related('created_by', 'updated_by', 'archived_by')
        
        file_access_archived = FileAccessNetworkIncident.objects.filter(
            is_archived=True
        ).select_related('created_by', 'updated_by', 'archived_by')
        
        radio_access_archived = RadioAccessNetworkIncident.objects.filter(
            is_archived=True
        ).select_related('created_by', 'updated_by', 'archived_by')
        
        core_archived = CoreNetworkIncident.objects.filter(
            is_archived=True
        ).select_related('created_by', 'updated_by', 'archived_by')
        
        backbone_archived = BackboneInternetNetworkIncident.objects.filter(
            is_archived=True
        ).select_related('created_by', 'updated_by', 'archived_by')
        
        # Combine all querysets and add network_type attribute
        all_archived = []
        
        for incident in transport_archived:
            incident.network_type_key = 'transport'
            incident.network_display_name = 'Transport Networks'
            all_archived.append(incident)
        
        for incident in file_access_archived:
            incident.network_type_key = 'file_access'
            incident.network_display_name = 'File Access Networks'
            all_archived.append(incident)
        
        for incident in radio_access_archived:
            incident.network_type_key = 'radio_access'
            incident.network_display_name = 'Radio Access Networks'
            all_archived.append(incident)
        
        for incident in core_archived:
            incident.network_type_key = 'core'
            incident.network_display_name = 'Core Networks'
            all_archived.append(incident)
        
        for incident in backbone_archived:
            incident.network_type_key = 'backbone_internet'
            incident.network_display_name = 'Backbone Internet Networks'
            all_archived.append(incident)
        
        # Apply filters from GET parameters
        network_filter = request.GET.get('network_type', '')
        cause_filter = request.GET.get('cause', '')
        origin_filter = request.GET.get('origin', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        search_query = request.GET.get('search', '')
        
        # Filter by network type
        if network_filter:
            all_archived = [inc for inc in all_archived if inc.network_type_key == network_filter]
        
        # Filter by cause
        if cause_filter:
            all_archived = [inc for inc in all_archived if inc.cause and cause_filter.lower() in inc.cause.lower()]
        
        # Filter by origin
        if origin_filter:
            all_archived = [inc for inc in all_archived if inc.origin and origin_filter.lower() in inc.origin.lower()]
        
        # Filter by date range
        if date_from:
            try:
                from datetime import datetime
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                date_from_aware = timezone.make_aware(date_from_obj) if timezone.is_naive(date_from_obj) else date_from_obj
                all_archived = [inc for inc in all_archived if inc.archived_at and inc.archived_at >= date_from_aware]
            except ValueError:
                pass
        
        if date_to:
            try:
                from datetime import datetime, time
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                # Set time to end of day
                date_to_obj = datetime.combine(date_to_obj.date(), time.max)
                date_to_aware = timezone.make_aware(date_to_obj) if timezone.is_naive(date_to_obj) else date_to_obj
                all_archived = [inc for inc in all_archived if inc.archived_at and inc.archived_at <= date_to_aware]
            except ValueError:
                pass
        
        # Search filter
        if search_query:
            search_lower = search_query.lower()
            all_archived = [inc for inc in all_archived 
                          if (str(inc.id).lower().startswith(search_lower) or
                              (inc.impact_comment and search_lower in inc.impact_comment.lower()) or
                              (inc.cause and search_lower in inc.cause.lower()) or
                              (inc.origin and search_lower in inc.origin.lower()))]
        
        # Sort by archived_at (newest first)
        all_archived.sort(key=lambda x: x.archived_at if x.archived_at else timezone.now(), reverse=True)
        
        # Calculate statistics by network
        stats_by_network = {
            'transport': len([i for i in all_archived if i.network_type_key == 'transport']),
            'file_access': len([i for i in all_archived if i.network_type_key == 'file_access']),
            'radio_access': len([i for i in all_archived if i.network_type_key == 'radio_access']),
            'core': len([i for i in all_archived if i.network_type_key == 'core']),
            'backbone_internet': len([i for i in all_archived if i.network_type_key == 'backbone_internet']),
        }
        
        # Get unique causes and origins for filter dropdowns
        all_causes = set()
        all_origins = set()
        for incident in all_archived:
            if incident.cause:
                all_causes.add(incident.cause)
            if incident.origin:
                all_origins.add(incident.origin)
        
        # Pagination
        paginator = Paginator(all_archived, 25)
        page_number = request.GET.get('page', 1)
        
        try:
            incidents_page = paginator.page(page_number)
        except (EmptyPage, InvalidPage):
            incidents_page = paginator.page(1)
        
        context = {
            'incidents': incidents_page,
            'page_obj': incidents_page,
            'total_archived': len(all_archived),
            'stats_by_network': stats_by_network,
            'all_causes': sorted(list(all_causes)),
            'all_origins': sorted(list(all_origins)),
            'current_filters': {
                'network_type': network_filter,
                'cause': cause_filter,
                'origin': origin_filter,
                'date_from': date_from,
                'date_to': date_to,
                'search': search_query,
            },
            'filter_active': any([network_filter, cause_filter, origin_filter, date_from, date_to, search_query]),
        }
        
        return render(request, 'incidents/unified_historical.html', context)
        
    except Exception as e:
        messages.error(request, f"Error loading historical incidents: {str(e)}")
        return render(request, 'incidents/unified_historical.html', {
            'incidents': [],
            'total_archived': 0,
            'stats_by_network': {},
            'all_causes': [],
            'all_origins': [],
            'current_filters': {},
            'filter_active': False,
            'error': str(e)
        })


@login_required
@require_http_methods(["POST"])
def restore_archived_incident(request, incident_id, network_type):
    """
    Restore an archived incident back to active status.
    Only admins can restore incidents.
    """
    # Check admin permission
    if not request.user.is_admin():
        messages.error(request, "Only administrators can restore archived incidents.")
        return redirect('incidents:unified_historical')
    
    try:
        # Get the appropriate model
        model_map = {
            'transport': TransportNetworkIncident,
            'file_access': FileAccessNetworkIncident,
            'radio_access': RadioAccessNetworkIncident,
            'core': CoreNetworkIncident,
            'backbone_internet': BackboneInternetNetworkIncident,
        }
        
        if network_type not in model_map:
            messages.error(request, f"Invalid network type: {network_type}")
            return redirect('incidents:unified_historical')
        
        model = model_map[network_type]
        incident = get_object_or_404(model, id=incident_id)
        
        # Check if incident is archived
        if not incident.is_archived:
            messages.warning(request, f"Incident {str(incident_id)[:8]} is not archived.")
            return redirect('incidents:unified_historical')
        
        # Restore the incident
        success = incident.restore(request.user)
        
        if success:
            messages.success(
                request, 
                f"Incident {str(incident_id)[:8]} successfully restored from archive."
            )
        else:
            messages.error(
                request,
                f"Failed to restore incident {str(incident_id)[:8]}. Please try again."
            )
        
    except Exception as e:
        messages.error(request, f"Error restoring incident: {str(e)}")
    
    return redirect('incidents:unified_historical')

@login_required
@require_http_methods(["POST"])
def archive_incident_manual(request, incident_id, network_type):
    """
    Manually archive an incident (Admin only).
    Validates archival criteria before archiving.
    """
    # Check admin permission
    if not request.user.is_admin():
        return JsonResponse({
            'success': False,
            'error': 'Only administrators can archive incidents manually.'
        }, status=403)
    
    try:
        # Get the appropriate model
        model_map = {
            'transport': TransportNetworkIncident,
            'file_access': FileAccessNetworkIncident,
            'radio_access': RadioAccessNetworkIncident,
            'core': CoreNetworkIncident,
            'backbone_internet': BackboneInternetNetworkIncident,
        }
        
        if network_type not in model_map:
            return JsonResponse({
                'success': False,
                'error': f'Invalid network type: {network_type}'
            }, status=400)
        
        model = model_map[network_type]
        incident = get_object_or_404(model, id=incident_id)
        
        # Check if incident can be archived
        if not incident.can_be_archived():
            return JsonResponse({
                'success': False,
                'error': 'Incident does not meet archival criteria. Ensure it is resolved, has cause and origin filled, and 2+ hours have passed since resolution.'
            }, status=400)
        
        # Archive the incident
        success = incident.archive(request.user)
        
        if success:
            return JsonResponse({
                'success': True,
                'message': f'Incident {str(incident_id)[:8]} successfully archived.',
                'incident_id': str(incident_id)
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Failed to archive incident. Please try again.'
            }, status=500)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
@login_required
@require_http_methods(["POST"])
def bulk_archive_incidents(request):
    """
    Bulk archive multiple incidents (Admin only).
    Receives list of incident IDs and archives all eligible ones.
    """
    # Check admin permission
    if not request.user.is_admin():
        return JsonResponse({
            'success': False,
            'error': 'Only administrators can bulk archive incidents.'
        }, status=403)
    
    try:
        import json
        data = json.loads(request.body)
        incident_ids = data.get('incident_ids', [])
        network_type = data.get('network_type')
        
        if not incident_ids or not network_type:
            return JsonResponse({
                'success': False,
                'error': 'Missing incident IDs or network type.'
            }, status=400)
        
        # Get the appropriate model
        model_map = {
            'transport': TransportNetworkIncident,
            'file_access': FileAccessNetworkIncident,
            'radio_access': RadioAccessNetworkIncident,
            'core': CoreNetworkIncident,
            'backbone_internet': BackboneInternetNetworkIncident,
        }
        
        if network_type not in model_map:
            return JsonResponse({
                'success': False,
                'error': f'Invalid network type: {network_type}'
            }, status=400)
        
        model = model_map[network_type]
        
        # Archive each incident
        archived_count = 0
        failed_incidents = []
        
        for incident_id in incident_ids:
            try:
                incident = model.objects.get(id=incident_id)
                
                # Admin can bulk archive any resolved incident (bypass 2-hour rule)
                if incident.date_time_recovery:
                    success = incident.archive(request.user)
                    if success:
                        archived_count += 1
                    else:
                        failed_incidents.append(str(incident_id)[:8])
                else:
                    # Not resolved yet
                    failed_incidents.append(str(incident_id)[:8])
                    
            except model.DoesNotExist:
                failed_incidents.append(str(incident_id)[:8])
        
        response_data = {
            'success': True,
            'archived_count': archived_count,
            'total_requested': len(incident_ids)
        }
        
        if failed_incidents:
            response_data['failed_incidents'] = failed_incidents
            response_data['message'] = f'Archived {archived_count} of {len(incident_ids)} incidents. Some incidents were not eligible.'
        else:
            response_data['message'] = f'Successfully archived {archived_count} incidents.'
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
# ============================================================================
# SAVED SEARCH FUNCTIONALITY (Task 1: Phase 4)
# ============================================================================

@login_required
@require_http_methods(["POST"])
def save_search_view(request, network_type):
    """
    Save current search filters for quick access later.
    Each user can have multiple saved searches per network type.
    """
    from .models import SavedSearch
    
    if network_type not in NETWORKS:
        return JsonResponse({
            'success': False,
            'error': 'Invalid network type'
        }, status=400)
    
    try:
        import json
        data = json.loads(request.body)
        
        search_name = data.get('name', '').strip()
        search_params = data.get('params', {})
        description = data.get('description', '').strip()
        set_as_default = data.get('set_as_default', False)
        
        # Validation
        if not search_name:
            return JsonResponse({
                'success': False,
                'error': 'Search name is required'
            }, status=400)
        
        if len(search_name) > 100:
            return JsonResponse({
                'success': False,
                'error': 'Search name must be 100 characters or less'
            }, status=400)
        
        # Check if user already has a saved search with this name for this network
        existing = SavedSearch.objects.filter(
            user=request.user,
            name=search_name,
            network_type=network_type
        ).first()
        
        if existing:
            # Update existing saved search
            existing.search_params = search_params
            existing.description = description
            existing.is_default = set_as_default
            existing.save()
            saved_search = existing
            created = False
        else:
            # Create new saved search
            saved_search = SavedSearch.objects.create(
                user=request.user,
                name=search_name,
                network_type=network_type,
                search_params=search_params,
                description=description,
                is_default=set_as_default
            )
            created = True
        
        # If setting as default, unset other defaults for this network
        if set_as_default:
            SavedSearch.objects.filter(
                user=request.user,
                network_type=network_type,
                is_default=True
            ).exclude(id=saved_search.id).update(is_default=False)
        
        return JsonResponse({
            'success': True,
            'message': f"Search '{search_name}' {'updated' if not created else 'saved'} successfully",
            'search_id': str(saved_search.id),
            'created': created
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["GET"])
def list_saved_searches_view(request, network_type):
    """
    Get all saved searches for the current user and network type.
    Returns JSON array of saved searches with metadata.
    """
    from .models import SavedSearch
    
    if network_type not in NETWORKS:
        return JsonResponse({
            'success': False,
            'error': 'Invalid network type'
        }, status=400)
    
    try:
        saved_searches = SavedSearch.objects.filter(
            user=request.user,
            network_type=network_type
        ).order_by('-last_used_at', '-created_at')
        
        searches_data = []
        for search in saved_searches:
            searches_data.append({
                'id': str(search.id),
                'name': search.name,
                'description': search.description,
                'params': search.search_params,
                'params_summary': search.get_params_summary(),
                'is_default': search.is_default,
                'use_count': search.use_count,
                'last_used_at': search.last_used_at.isoformat() if search.last_used_at else None,
                'created_at': search.created_at.isoformat()
            })
        
        return JsonResponse({
            'success': True,
            'searches': searches_data,
            'count': len(searches_data)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def load_saved_search_view(request, search_id):
    """
    Load a saved search by ID and increment its usage counter.
    Returns the search parameters to populate the search form.
    """
    from .models import SavedSearch
    
    try:
        saved_search = get_object_or_404(
            SavedSearch,
            id=search_id,
            user=request.user  # Security: Only load user's own searches
        )
        
        # Increment usage counter
        saved_search.increment_use_count()
        
        return JsonResponse({
            'success': True,
            'search': {
                'id': str(saved_search.id),
                'name': saved_search.name,
                'description': saved_search.description,
                'params': saved_search.search_params,
                'network_type': saved_search.network_type,
                'is_default': saved_search.is_default
            }
        })
        
    except SavedSearch.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Saved search not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def delete_saved_search_view(request, search_id):
    """
    Delete a saved search by ID.
    Only the owner can delete their saved searches.
    """
    from .models import SavedSearch
    
    try:
        saved_search = get_object_or_404(
            SavedSearch,
            id=search_id,
            user=request.user  # Security: Only delete user's own searches
        )
        
        search_name = saved_search.name
        saved_search.delete()
        
        return JsonResponse({
            'success': True,
            'message': f"Search '{search_name}' deleted successfully"
        })
        
    except SavedSearch.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Saved search not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def set_default_search_view(request, search_id):
    """
    Set a saved search as the default for its network type.
    Unsets any other default for that network.
    """
    from .models import SavedSearch
    
    try:
        saved_search = get_object_or_404(
            SavedSearch,
            id=search_id,
            user=request.user  # Security: Only modify user's own searches
        )
        
        # Unset other defaults for this network type
        SavedSearch.objects.filter(
            user=request.user,
            network_type=saved_search.network_type,
            is_default=True
        ).exclude(id=saved_search.id).update(is_default=False)
        
        # Set this search as default
        saved_search.is_default = True
        saved_search.save(update_fields=['is_default'])
        
        return JsonResponse({
            'success': True,
            'message': f"'{saved_search.name}' set as default search"
        })
        
    except SavedSearch.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Saved search not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
# ============================================================================
# CSV/EXCEL EXPORT FUNCTIONALITY (Task 2: Phase 4)
# ============================================================================

from django.http import HttpResponse
from .services import get_export_service


@login_required
@require_http_methods(["POST"])
def export_incidents_view(request, network_type):
    """
    Export filtered incidents to CSV or Excel format.
    Respects current search filters and exports only what user sees.
    """
    if network_type not in NETWORKS:
        return JsonResponse({
            'success': False,
            'error': 'Invalid network type'
        }, status=400)
    
    try:
        import json
        data = json.loads(request.body)
        
        export_format = data.get('format', 'csv').lower()
        search_params = data.get('search_params', {})
        
        # Validate format
        if export_format not in ['csv', 'xlsx', 'excel']:
            return JsonResponse({
                'success': False,
                'error': 'Invalid export format. Use "csv" or "xlsx"'
            }, status=400)
        
        # Normalize excel format
        if export_format == 'excel':
            export_format = 'xlsx'
        
        # Get the model
        model = NETWORKS[network_type]['model']
        
        # Build queryset with same filters as current view
        queryset = model.objects.filter(is_archived=False).select_related('created_by', 'updated_by')
        
        # Apply search filters if provided
        if search_params:
            search_service = get_search_service(network_type)
            queryset = search_service.search_incidents(search_params)
        else:
            queryset = queryset.order_by('-date_time_incident')
        
        # Limit to prevent memory issues (configurable)
        max_export_limit = 10000
        total_count = queryset.count()
        
        if total_count > max_export_limit:
            return JsonResponse({
                'success': False,
                'error': f'Export limit exceeded. Maximum {max_export_limit} incidents allowed. Your filter returned {total_count} incidents. Please narrow your search.'
            }, status=400)
        
        # Get export service
        export_service = get_export_service(queryset, network_type)
        
        # Generate export file
        if export_format == 'csv':
            content = export_service.export_to_csv()
            content_type = 'text/csv'
        else:  # xlsx
            content = export_service.export_to_excel()
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Get filename
        filename = export_service.get_filename(export_format)
        
        # Create response
        response = HttpResponse(content, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Log export activity (optional)
        from .models import AuditLog
        AuditLog.objects.create(
            user=request.user,
            action='EXPORT',
            model_name=f'{network_type}_incidents',
            object_id=f'{total_count}_records',
            changes={'format': export_format, 'count': total_count},
            ip_address=request.META.get('REMOTE_ADDR'),
        )
        
        return response
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Export failed: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def bulk_export_all_networks(request):
    """
    Export incidents from all networks into a single Excel file with multiple sheets.
    Admin-only feature for comprehensive reporting.
    """
    # Check admin permission
    if not request.user.is_admin():
        return JsonResponse({
            'success': False,
            'error': 'Only administrators can export all networks'
        }, status=403)
    
    try:
        import json
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        data = json.loads(request.body)
        date_from = data.get('date_from')
        date_to = data.get('date_to')
        include_archived = data.get('include_archived', False)
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        total_incidents = 0
        
        # Export each network type to separate sheet
        for network_key, network_config in NETWORKS.items():
            model = network_config['model']
            
            # Build queryset
            queryset = model.objects.select_related('created_by', 'updated_by')
            
            if not include_archived:
                queryset = queryset.filter(is_archived=False)
            
            # Apply date filters
            if date_from:
                queryset = queryset.filter(date_time_incident__gte=date_from)
            if date_to:
                queryset = queryset.filter(date_time_incident__lte=date_to)
            
            queryset = queryset.order_by('-date_time_incident')
            
            # Get export service
            export_service = get_export_service(queryset, network_key)
            
            # Create sheet
            sheet_name = network_config['name'][:31]  # Excel limit
            ws = wb.create_sheet(sheet_name)
            
            # Get headers and data
            headers, field_getters = export_service._get_export_fields()
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="003d7a", end_color="003d7a", fill_type="solid")
            
            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Write data
            for row_num, incident in enumerate(queryset, 2):
                for col_num, getter in enumerate(field_getters, 1):
                    ws.cell(row=row_num, column=col_num, value=getter(incident))
            
            total_incidents += queryset.count()
        
        # Create summary sheet at the beginning
        summary_ws = wb.create_sheet('Summary', 0)
        summary_ws['A1'] = 'Network Incident Export Summary'
        summary_ws['A1'].font = Font(bold=True, size=14)
        
        summary_ws['A3'] = 'Export Date:'
        summary_ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        summary_ws['A4'] = 'Exported By:'
        summary_ws['B4'] = request.user.username
        
        summary_ws['A5'] = 'Total Incidents:'
        summary_ws['B5'] = total_incidents
        
        if date_from:
            summary_ws['A6'] = 'Date From:'
            summary_ws['B6'] = date_from
        
        if date_to:
            summary_ws['A7'] = 'Date To:'
            summary_ws['B7'] = date_to
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'all_networks_export_{timestamp}.xlsx'
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Log export
        from .models import AuditLog
        AuditLog.objects.create(
            user=request.user,
            action='EXPORT',
            model_name='all_networks',
            object_id=f'{total_incidents}_records',
            changes={'format': 'xlsx', 'count': total_incidents},
            ip_address=request.META.get('REMOTE_ADDR'),
        )
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Bulk export failed: {str(e)}'
        }, status=500)